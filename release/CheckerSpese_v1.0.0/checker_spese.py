#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Bot per la pulizia e validazione dei dati delle spese
"""

import os
import re
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
from typing import List, Tuple, Dict


class CheckerSpese:
    """Classe principale per il controllo e pulizia delle spese"""

    # Mapping colonne (1-based come in Excel)
    COLS = {
        'CODICE_ATTIVITA': 1,      # A
        'CODPAG': 2,               # B
        'PROGETTO': 3,             # C
        'CUP': 4,                  # D
        'SOGGETTO': 5,             # E
        'TIPOLOGIA_SPESA': 9,      # I
        'INQUADRAMENTO': 19,       # S
        'TIPOLOGIA_REND': 21,      # U
        'DESCRIZIONE_VOCE': 22,    # V
        'IMPORTO_TOTALE': 26,      # Z
        'STATO': 46,               # AT
    }

    # Dipartimenti validi
    DIPARTIMENTI = [
        'DAER', 'DCMC', 'DEIB', 'DENG', 'DICA', 'DIG_',
        'DMAT', 'DMEC', 'DASTU', 'DFIS', 'DESIGN', 'DABC'
    ]

    # Pattern per inquadramenti validi
    INQUADRAMENTI_VALIDI = [
        'Ordinario', 'Associato', 'Ricercatore', 'RTD', 'PO', 'PA'
    ]

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = Path(file_path).stem
        self.modifiche = []
        self.errori_rows = []
        self.wb = None
        self.ws = None
        self.righe_eliminate = 0

    def log_modifica(self, messaggio: str):
        """Registra una modifica nel log"""
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.modifiche.append(f"[{timestamp}] {messaggio}")
        print(f"  → {messaggio}")

    def carica_file(self):
        """Carica il file Excel"""
        print(f"Caricamento file: {self.file_path}")
        self.wb = openpyxl.load_workbook(self.file_path)
        self.ws = self.wb.active
        self.log_modifica(f"File caricato: {self.file_path}")
        self.log_modifica(f"Totale righe iniziali: {self.ws.max_row - 1}")

    def fase1_elimina_non_polimi(self):
        """Fase 1: Elimina righe dove Soggetto non contiene POLIMI"""
        print("\n=== FASE 1: Eliminazione spese non POLIMI ===")
        col_soggetto = self.COLS['SOGGETTO']
        righe_da_eliminare = []

        # Scansiona dal basso verso l'alto per evitare problemi con l'eliminazione
        for row in range(self.ws.max_row, 1, -1):
            soggetto = self.ws.cell(row, col_soggetto).value
            if soggetto and 'POLIMI' not in str(soggetto).upper():
                righe_da_eliminare.append(row)

        # Elimina le righe
        for row in righe_da_eliminare:
            self.ws.delete_rows(row)
            self.righe_eliminate += 1

        self.log_modifica(f"Fase 1: Eliminate {len(righe_da_eliminare)} righe non POLIMI")

    def fase2_elimina_stati_non_validi(self):
        """Fase 2: Elimina righe con stati diversi da Trasmessa o Conclusa in attesa"""
        print("\n=== FASE 2: Eliminazione stati non validi ===")
        col_stato = self.COLS['STATO']
        stati_validi = ['TRASMESSA', 'CONCLUSA IN ATTESA TRASMISSIONE ATTESTAZIONE']
        righe_da_eliminare = []

        for row in range(self.ws.max_row, 1, -1):
            stato = self.ws.cell(row, col_stato).value
            if stato and str(stato).upper().strip() not in stati_validi:
                righe_da_eliminare.append(row)

        for row in righe_da_eliminare:
            self.ws.delete_rows(row)
            self.righe_eliminate += 1

        self.log_modifica(f"Fase 2: Eliminate {len(righe_da_eliminare)} righe con stato non valido")

    def fase3_elimina_costi_indiretti(self):
        """Fase 3: Elimina righe con Tipologia spesa = Costi indiretti"""
        print("\n=== FASE 3: Eliminazione costi indiretti ===")
        col_tipo_spesa = self.COLS['TIPOLOGIA_SPESA']
        righe_da_eliminare = []

        for row in range(self.ws.max_row, 1, -1):
            tipo_spesa = self.ws.cell(row, col_tipo_spesa).value
            if tipo_spesa and str(tipo_spesa).upper().strip() == 'COSTI INDIRETTI':
                righe_da_eliminare.append(row)

        for row in righe_da_eliminare:
            self.ws.delete_rows(row)
            self.righe_eliminate += 1

        self.log_modifica(f"Fase 3: Eliminate {len(righe_da_eliminare)} righe con costi indiretti")

    def fase4_pulizia_dipartimenti(self):
        """Fase 4: Pulizia e correzione dei dipartimenti"""
        print("\n=== FASE 4: Pulizia dipartimenti ===")
        col_tipo_spesa = self.COLS['TIPOLOGIA_SPESA']
        col_descrizione = self.COLS['DESCRIZIONE_VOCE']
        col_codpag = self.COLS['CODPAG']

        modifiche_auto = 0
        righe_da_verificare = []

        # Scansiona tutte le righe (escluso header)
        for row in range(2, self.ws.max_row + 1):
            tipo_spesa = self.ws.cell(row, col_tipo_spesa).value

            # Salta "Erogazione bandi a cascata"
            if tipo_spesa and 'EROGAZIONE BANDI A CASCATA' in str(tipo_spesa).upper():
                continue

            descrizione = self.ws.cell(row, col_descrizione).value
            if not descrizione:
                continue

            descrizione_str = str(descrizione).strip()
            codpag = self.ws.cell(row, col_codpag).value

            # Step 1: Verifica se inizia con un dipartimento valido
            trovato = False
            for dip in self.DIPARTIMENTI:
                if descrizione_str.upper().startswith(dip.upper()):
                    trovato = True
                    break

            if trovato:
                continue

            # Step 2: Prova a correggere errori comuni con regex
            correzione = self._correggi_dipartimento(descrizione_str)
            if correzione and correzione != descrizione_str:
                self.ws.cell(row, col_descrizione).value = correzione
                self.log_modifica(f"Riga {row} (CODPAG {codpag}): Corretto '{descrizione_str[:50]}...' -> '{correzione[:50]}...'")
                modifiche_auto += 1
                continue

            # Step 3: Cerca occorrenze di dipartimenti nel testo
            dip_trovato = None
            for dip in self.DIPARTIMENTI:
                if re.search(r'\b' + re.escape(dip) + r'\b', descrizione_str, re.IGNORECASE):
                    dip_trovato = dip
                    break

            if dip_trovato:
                proposta = f"{dip_trovato}_{descrizione_str}"
                righe_da_verificare.append({
                    'row': row,
                    'codpag': codpag,
                    'originale': descrizione_str,
                    'proposta': proposta,
                    'dipartimento': dip_trovato
                })
            else:
                # Step 4: Aggiungi a errori
                self._aggiungi_errore(row, f"Dipartimento non riconosciuto in descrizione")

        self.log_modifica(f"Fase 4: Effettuate {modifiche_auto} correzioni automatiche")

        # Mostra le righe da verificare all'utente
        if righe_da_verificare:
            self._mostra_modal_verifiche_dipartimenti(righe_da_verificare)

    def _correggi_dipartimento(self, testo: str) -> str:
        """Applica correzioni automatiche ai dipartimenti"""
        testo_originale = testo

        # Rimuovi spazi iniziali
        testo = testo.lstrip()

        # Rimuovi "POLIMI" o "POLI" iniziale
        testo = re.sub(r'^(POLIMI[-_\s]+|POLI[-_\s]+)', '', testo, flags=re.IGNORECASE)

        # Correzioni comuni
        correzioni = {
            r'^DIG\.': 'DIG_',
            r'^CMC\b': 'DCMC',
            r'^POLI\b': 'DEIB',
            r'^DESING\b': 'DESIGN',
            r'^DESIGNN\b': 'DESIGN',
            r'^DESGN\b': 'DESIGN',
            r'^DEIBB\b': 'DEIB',
            r'^DEIB\s*-': 'DEIB_',
        }

        for pattern, sostituzione in correzioni.items():
            testo = re.sub(pattern, sostituzione, testo, flags=re.IGNORECASE)

        # Se abbiamo fatto modifiche, assicurati che inizi con un dipartimento valido
        for dip in self.DIPARTIMENTI:
            if testo.upper().startswith(dip.upper()):
                # Normalizza il caso
                testo = dip + testo[len(dip):]
                break

        return testo if testo != testo_originale else testo_originale

    def _mostra_modal_verifiche_dipartimenti(self, righe: List[Dict]):
        """Mostra un modal per la verifica delle correzioni proposte"""
        root = tk.Tk()
        root.title("Verifiche dipartimenti da confermare")
        root.geometry("900x600")

        # Frame principale
        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Label info
        ttk.Label(main_frame, text=f"Trovate {len(righe)} righe con dipartimenti da verificare",
                  font=('Arial', 12, 'bold')).grid(row=0, column=0, columnspan=3, pady=10)

        # Treeview per mostrare le righe
        columns = ('CODPAG', 'Originale', 'Proposta', 'Dipartimento')
        tree = ttk.Treeview(main_frame, columns=columns, show='tree headings', height=20)

        tree.heading('#0', text='Applica')
        tree.heading('CODPAG', text='CODPAG')
        tree.heading('Originale', text='Originale')
        tree.heading('Proposta', text='Proposta')
        tree.heading('Dipartimento', text='Dip.')

        tree.column('#0', width=50)
        tree.column('CODPAG', width=100)
        tree.column('Originale', width=300)
        tree.column('Proposta', width=300)
        tree.column('Dipartimento', width=80)

        # Scrollbar
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=1, column=3, sticky=(tk.N, tk.S))

        # Popola il tree
        for riga in righe:
            tree.insert('', tk.END, text='☐',
                       values=(riga['codpag'],
                              riga['originale'][:50] + '...' if len(riga['originale']) > 50 else riga['originale'],
                              riga['proposta'][:50] + '...' if len(riga['proposta']) > 50 else riga['proposta'],
                              riga['dipartimento']),
                       tags=('unchecked',))

        # Gestione click per toggle checkbox
        def toggle_check(event):
            item = tree.selection()[0]
            current_text = tree.item(item, 'text')
            new_text = '☑' if current_text == '☐' else '☐'
            tree.item(item, text=new_text)

        tree.bind('<Button-1>', toggle_check)

        # Pulsanti
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=2, column=0, columnspan=3, pady=10)

        def seleziona_tutti():
            for item in tree.get_children():
                tree.item(item, text='☑')

        def deseleziona_tutti():
            for item in tree.get_children():
                tree.item(item, text='☐')

        def applica_modifiche():
            modifiche_applicate = 0
            for i, item in enumerate(tree.get_children()):
                if tree.item(item, 'text') == '☑':
                    riga_data = righe[i]
                    row = riga_data['row']
                    col_descrizione = self.COLS['DESCRIZIONE_VOCE']
                    self.ws.cell(row, col_descrizione).value = riga_data['proposta']
                    self.log_modifica(f"Riga {row} (CODPAG {riga_data['codpag']}): Applicata correzione manuale")
                    modifiche_applicate += 1
                else:
                    # Non applicata, aggiungi a errori
                    riga_data = righe[i]
                    self._aggiungi_errore(riga_data['row'], "Correzione dipartimento non confermata dall'utente")

            messagebox.showinfo("Completato", f"Applicate {modifiche_applicate} modifiche")
            root.destroy()

        ttk.Button(button_frame, text="Seleziona tutti", command=seleziona_tutti).grid(row=0, column=0, padx=5)
        ttk.Button(button_frame, text="Deseleziona tutti", command=deseleziona_tutti).grid(row=0, column=1, padx=5)
        ttk.Button(button_frame, text="Applica modifiche", command=applica_modifiche).grid(row=0, column=2, padx=5)
        ttk.Button(button_frame, text="Salta tutto", command=lambda: [self._aggiungi_errori_batch(righe), root.destroy()]).grid(row=0, column=3, padx=5)

        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)

        root.mainloop()

    def _aggiungi_errori_batch(self, righe: List[Dict]):
        """Aggiunge un batch di righe agli errori"""
        for riga in righe:
            self._aggiungi_errore(riga['row'], "Correzione dipartimento non confermata")

    def fase5_validazione_rendicontazione(self):
        """Fase 5: Validazione delle regole di rendicontazione"""
        print("\n=== FASE 5: Validazione rendicontazione ===")
        col_tipo_spesa = self.COLS['TIPOLOGIA_SPESA']
        col_inquadramento = self.COLS['INQUADRAMENTO']
        col_tipo_rend = self.COLS['TIPOLOGIA_REND']
        col_codpag = self.COLS['CODPAG']

        errori_trovati = []

        for row in range(2, self.ws.max_row + 1):
            tipo_spesa = self.ws.cell(row, col_tipo_spesa).value
            inquadramento = self.ws.cell(row, col_inquadramento).value
            tipo_rend = self.ws.cell(row, col_tipo_rend).value
            codpag = self.ws.cell(row, col_codpag).value

            if not tipo_spesa:
                continue

            tipo_spesa_str = str(tipo_spesa).strip()
            inquadramento_str = str(inquadramento).strip() if inquadramento else ""
            tipo_rend_str = str(tipo_rend).strip() if tipo_rend else ""

            # Verifica se inquadramento è valido
            inquadramento_valido = self._is_inquadramento_valido(inquadramento_str)

            errore = None

            if 'SPESE DI PERSONALE' in tipo_spesa_str.upper():
                if inquadramento_valido:
                    # Deve essere a costi standard
                    if 'COSTI STANDARD' not in tipo_rend_str.upper():
                        errore = f"Spese personale con inquadramento valido deve avere rendicontazione a costi standard"
                else:
                    # Deve essere a costi reali
                    if 'COSTI REALI' not in tipo_rend_str.upper():
                        errore = f"Spese personale senza inquadramento valido deve avere rendicontazione a costi reali"

            elif any(x in tipo_spesa_str.upper() for x in ['ALTRE TIPOLOGIE', 'CONSULENZA', 'MATERIALI', 'ATTREZZATURE', 'LICENZE']):
                # Deve essere a costi reali
                if 'COSTI REALI' not in tipo_rend_str.upper():
                    errore = f"Altre spese devono avere rendicontazione a costi reali"
                # Verifica che inquadramento sia vuoto o non valido
                if inquadramento_valido:
                    errore = f"Altre spese non devono avere inquadramento valido"

            if errore:
                errori_trovati.append({
                    'row': row,
                    'codpag': codpag,
                    'tipo_spesa': tipo_spesa_str,
                    'inquadramento': inquadramento_str,
                    'tipo_rend': tipo_rend_str,
                    'errore': errore
                })

        self.log_modifica(f"Fase 5: Trovati {len(errori_trovati)} errori di validazione")

        if errori_trovati:
            self._mostra_modal_errori_validazione(errori_trovati)

    def _is_inquadramento_valido(self, inquadramento: str) -> bool:
        """Verifica se un inquadramento è valido"""
        if not inquadramento:
            return False
        inq_upper = inquadramento.upper()
        for valido in self.INQUADRAMENTI_VALIDI:
            if valido.upper() in inq_upper:
                return True
        return False

    def _mostra_modal_errori_validazione(self, errori: List[Dict]):
        """Mostra un modal con gli errori di validazione trovati"""
        root = tk.Tk()
        root.title("Errori di validazione")
        root.geometry("1000x600")

        main_frame = ttk.Frame(root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        ttk.Label(main_frame, text=f"Trovati {len(errori)} errori di validazione",
                  font=('Arial', 12, 'bold')).grid(row=0, column=0, columnspan=4, pady=10)

        # Treeview
        columns = ('CODPAG', 'Tipo Spesa', 'Inquadramento', 'Tipo Rend.', 'Errore')
        tree = ttk.Treeview(main_frame, columns=columns, show='headings', height=20)

        tree.heading('CODPAG', text='CODPAG')
        tree.heading('Tipo Spesa', text='Tipo Spesa')
        tree.heading('Inquadramento', text='Inquadramento')
        tree.heading('Tipo Rend.', text='Tipo Rend.')
        tree.heading('Errore', text='Errore')

        tree.column('CODPAG', width=100)
        tree.column('Tipo Spesa', width=200)
        tree.column('Inquadramento', width=150)
        tree.column('Tipo Rend.', width=150)
        tree.column('Errore', width=350)

        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        tree.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=1, column=4, sticky=(tk.N, tk.S))

        for err in errori:
            tree.insert('', tk.END, values=(
                err['codpag'],
                err['tipo_spesa'][:30],
                err['inquadramento'][:25],
                err['tipo_rend'][:25],
                err['errore']
            ))
            # Aggiungi a errori
            self._aggiungi_errore(err['row'], err['errore'])

        ttk.Button(main_frame, text="OK - Aggiunti a file errori",
                  command=root.destroy).grid(row=2, column=0, columnspan=4, pady=10)

        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)

        root.mainloop()

    def _aggiungi_errore(self, row: int, motivo: str):
        """Aggiunge una riga agli errori"""
        riga_dati = []
        for col in range(1, self.ws.max_column + 1):
            riga_dati.append(self.ws.cell(row, col).value)
        riga_dati.append(motivo)  # Aggiungi motivo come ultima colonna
        self.errori_rows.append(riga_dati)
        self.log_modifica(f"Riga {row}: Aggiunta a errori - {motivo}")

    def salva_output(self):
        """Salva i file di output"""
        print("\n=== Salvataggio output ===")

        # Salva file pulito
        output_clean = f"clean_{self.file_name}.xlsx"
        self.wb.save(output_clean)
        self.log_modifica(f"Salvato file pulito: {output_clean}")
        print(f"✓ File pulito salvato: {output_clean}")

        # Salva log modifiche
        output_log = f"modifiche_effettuate_{self.file_name}.txt"
        with open(output_log, 'w', encoding='utf-8') as f:
            f.write("=" * 80 + "\n")
            f.write("LOG MODIFICHE CHECKER SPESE\n")
            f.write("=" * 80 + "\n\n")
            for modifica in self.modifiche:
                f.write(modifica + "\n")
        print(f"✓ Log modifiche salvato: {output_log}")

        # Salva errori se presenti
        if self.errori_rows:
            output_errori = "errori.xlsx"
            wb_errori = openpyxl.Workbook()
            ws_errori = wb_errori.active
            ws_errori.title = "Errori"

            # Copia header + aggiungi colonna motivo
            header = []
            for col in range(1, self.ws.max_column + 1):
                header.append(self.ws.cell(1, col).value)
            header.append("MOTIVO ERRORE")
            ws_errori.append(header)

            # Aggiungi righe errori
            for riga in self.errori_rows:
                ws_errori.append(riga)

            wb_errori.save(output_errori)
            self.log_modifica(f"Salvato file errori: {output_errori} ({len(self.errori_rows)} righe)")
            print(f"✓ File errori salvato: {output_errori} ({len(self.errori_rows)} righe)")

    def esegui(self):
        """Esegue tutte le fasi del processo"""
        try:
            self.carica_file()
            self.fase1_elimina_non_polimi()
            self.fase2_elimina_stati_non_validi()
            self.fase3_elimina_costi_indiretti()
            self.fase4_pulizia_dipartimenti()
            self.fase5_validazione_rendicontazione()
            self.salva_output()

            print("\n" + "=" * 80)
            print("✓ PROCESSO COMPLETATO CON SUCCESSO")
            print("=" * 80)
            print(f"Righe totali eliminate: {self.righe_eliminate}")
            print(f"Righe finali nel file pulito: {self.ws.max_row - 1}")
            print(f"Righe con errori: {len(self.errori_rows)}")

            messagebox.showinfo("Completato",
                              f"Processo completato!\n\n"
                              f"Righe eliminate: {self.righe_eliminate}\n"
                              f"Righe finali: {self.ws.max_row - 1}\n"
                              f"Righe con errori: {len(self.errori_rows)}")

        except Exception as e:
            print(f"\n❌ ERRORE: {e}")
            messagebox.showerror("Errore", f"Si è verificato un errore:\n\n{e}")
            raise


def main():
    """Funzione principale"""
    print("=" * 80)
    print("CHECKER SPESE - Bot per pulizia dati")
    print("=" * 80)

    # Cerca file .xlsx nella directory corrente
    xlsx_files = [f for f in os.listdir('.') if f.endswith('.xlsx') and not f.startswith('clean_') and f != 'errori.xlsx']

    if not xlsx_files:
        print("❌ Nessun file .xlsx trovato nella directory corrente!")
        messagebox.showerror("Errore", "Nessun file .xlsx trovato nella directory corrente!")
        return

    if len(xlsx_files) == 1:
        file_path = xlsx_files[0]
        print(f"\nFile selezionato: {file_path}")
    else:
        print("\nFile .xlsx trovati:")
        for i, f in enumerate(xlsx_files, 1):
            print(f"  {i}. {f}")

        scelta = input(f"\nSeleziona il file (1-{len(xlsx_files)}): ").strip()
        try:
            idx = int(scelta) - 1
            if 0 <= idx < len(xlsx_files):
                file_path = xlsx_files[idx]
            else:
                print("Selezione non valida!")
                return
        except ValueError:
            print("Selezione non valida!")
            return

    # Esegui il checker
    checker = CheckerSpese(file_path)
    checker.esegui()


if __name__ == "__main__":
    main()
